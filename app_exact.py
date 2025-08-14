
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from pathlib import Path
import os, re, csv, zipfile, sys, importlib
from PIL import Image, ImageTk

try:
    import openpyxl  # noqa: F401
except Exception:
    openpyxl = None
try:
    import xlrd  # noqa: F401
except Exception:
    xlrd = None

from db_helper import ensure_db, upsert_many_batched, replace_all, query

APP_DIR = Path(__file__).parent
DB_PATH = str(APP_DIR / "codebook.db")

COMMON_ENCODINGS = ["utf-8-sig","utf-8","gbk","gb18030","utf-16","utf-16le","utf-16be","latin1"]
COMMON_DELIMS = ["|","\t",",",";"," "]

def sniff_delimiter(text: str):
    try:
        dialect = csv.Sniffer().sniff(text, delimiters="|\t,; ")
        return dialect.delimiter
    except Exception:
        first = next((ln for ln in text.splitlines() if ln.strip()), "")
        best, bestn = None, 1
        for d in COMMON_DELIMS:
            n = len([c for c in first.split(d) if c != ""])
            if n > bestn:
                best, bestn = d, n
        return best or ","

def try_parse_txt(path: str):
    from io import StringIO
    data = Path(path).read_bytes()
    for enc in COMMON_ENCODINGS:
        try:
            s = data.decode(enc)
        except Exception:
            continue
        s = s.replace("\r\n","\n").replace("\r","\n")
        if s and s[0] == "\ufeff": s = s[1:]
        try:
            delim = sniff_delimiter(s)
        except Exception:
            delim = None
        def try_read(sep):
            sio = StringIO(s)
            return pd.read_csv(sio, sep=sep, header=None, dtype=str, engine="python",
                               quoting=3, on_bad_lines="skip", escapechar="\\").dropna(axis=1, how="all").dropna(axis=0, how="all")
        for sep in [delim, "|","\t",",",";","\s+"]:
            if not sep: continue
            try:
                df = try_read(sep)
                if df is not None and not df.empty: break
            except Exception:
                df = None
        else:
            df = None
        if df is not None and df.shape[1] == 1:
            col = df.columns[0]
            for sep in ["|","\t",",",";"]:
                parts = df[col].str.split(sep, expand=True)
                if parts.shape[1] >= 2: df = parts; break
            else:
                parts = df[col].str.split(r"\s+", expand=True)
                if parts.shape[1] >= 2: df = parts
        if df is not None and not df.empty: return df
    raise RuntimeError("无法解析 TXT：请另存为 CSV/Excel 再导入。")

def read_any(path: str):
    p = Path(path); ext = p.suffix.lower()
    if ext in [".xlsx",".xlsm",".xltx",".xltm"]:
        if not zipfile.is_zipfile(path):
            raise RuntimeError("扩展名为 .xlsx，但内容不是 Office Open XML（可能被错误改名）。请改回正确扩展名或另存为 .xlsx 再试。")
        return pd.read_excel(path, engine="openpyxl", dtype=str)
    if ext == ".xls":
        if xlrd is None or getattr(importlib.import_module('xlrd'), '__version__', '') != '1.2.0':
            raise RuntimeError("读取 .xls 需要 xlrd==1.2.0，请在“帮助→环境自检与修复”查看修复指引，或将文件另存为 .xlsx/CSV。")
        return pd.read_excel(path, engine="xlrd", dtype=str)
    if ext == ".csv":
        return pd.read_csv(path, dtype=str, engine="python", sep=None, on_bad_lines="skip")
    if ext in [".txt",".dat"]:
        return try_parse_txt(path)
    raise RuntimeError("不支持的文件类型，请转存为 CSV/Excel 后再导入。")

def export_text_xlsx(df: pd.DataFrame, path: str, *, include_header: bool = True):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title="Sheet1"
    start_row = 1
    if include_header:
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=j, value=str(col)); cell.number_format="@"
        start_row = 2
    for i, (_, row) in enumerate(df.iterrows(), start=start_row):
        for j, col in enumerate(df.columns, start=1):
            val = "" if pd.isna(row[col]) else str(row[col])
            cell = ws.cell(row=i, column=j, value=val); cell.number_format="@"
    wb.save(path)

# ---------------- 背景水印 ----------------
def _install_watermark(frame, img_path, opacity=0.08):
    if not Path(img_path).exists():
        return
    canvas = getattr(frame, "_bg_canvas", None)
    if canvas is None:
        canvas = tk.Canvas(frame, borderwidth=0, highlightthickness=0)
        canvas.place(x=0, y=0, relwidth=1, relheight=1)
        canvas.lower()
        frame._bg_canvas = canvas
    pil = getattr(frame, "_bg_pil", None)
    if pil is None or getattr(frame, "_bg_src", "") != str(img_path):
        try:
            pil = Image.open(img_path).convert("RGBA")
        except Exception:
            return
        frame._bg_pil = pil
        frame._bg_src = str(img_path)
    def _render(event=None):
        if not hasattr(frame, "_bg_pil"):
            return
        w = max(10, frame.winfo_width()); h = max(10, frame.winfo_height())
        if w < 10 or h < 10: return
        pil = frame._bg_pil
        scale = max(w / pil.width, h / pil.height)
        new_size = (int(pil.width * scale), int(pil.height * scale))
        im = pil.resize(new_size, Image.LANCZOS)
        left = (im.width - w) // 2; top = (im.height - h) // 2
        im = im.crop((max(0,left), max(0,top), max(w, left+w), max(h, top+h)))

        # 制作低透明淡化效果
        # 前景：原图，设置 alpha；背景：白色
        fg = im.copy(); fg.putalpha(int(255*opacity))
        bg = Image.new("RGBA", im.size, (255,255,255,255))
        im = Image.alpha_composite(bg, fg).convert("RGB")

        frame._bg_imgtk = ImageTk.PhotoImage(im)
        canvas.delete("all")
        canvas.create_image(0, 0, image=frame._bg_imgtk, anchor="nw")
    frame.bind("<Configure>", _render)
    frame.after(50, _render)

# ---------------- IBPS/CNAPS 解析 ----------------
def _locate_header_row_for_ibps(df):
    scan = df.head(30).astype(str).fillna("")
    k_code = ["清算行行号","清算行号","联行号","行号","行号代码","清算行行号代码"]
    k_name = ["清算行名称","清算行名","名称","银行名称","开户行名称"]
    for ridx in range(len(scan)):
        row = [str(x).strip() for x in scan.iloc[ridx].tolist()]
        joined = "".join(row)
        if not joined:
            continue
        has_code_kw = any(kw in joined for kw in k_code)
        has_name_kw = any(kw in joined for kw in k_name)
        non_empty_cols = sum(1 for x in row if x != "")
        if has_code_kw and has_name_kw and non_empty_cols >= 2:
            return ridx
    return None

def pick_ibps(df):
    df2 = df.copy()
    hdr = _locate_header_row_for_ibps(df2)
    if hdr is not None:
        df2.columns = [str(c).strip() for c in df2.iloc[hdr].tolist()]
        df2 = df2.iloc[hdr+1:].reset_index(drop=True)
    df2.columns = [str(c).strip() for c in df2.columns]
    code_alias = ["清算行行号","清算行号","联行号","行号","行号代码","清算行行号代码"]
    name_alias = ["清算行名称","清算行名","名称","银行名称","开户行名称"]
    code_col = next((c for c in df2.columns if c in code_alias), None)
    name_col = next((c for c in df2.columns if c in name_alias), None)
    if code_col and name_col:
        use = df2[[code_col, name_col]].copy(); use.columns = ["code","name"]
    else:
        use = df2.iloc[:, :2].copy(); use.columns = ["code","name"]
    use["code"] = use["code"].astype(str).str.replace(r"\.0$", "", regex=True)
    use["code"] = use["code"].str.extract(r"(\d{12})", expand=False).fillna("")
    use["name"] = use["name"].astype(str).str.replace("\u3000"," ").str.strip()
    use = use[(use["code"]!="") & (use["name"]!="")].drop_duplicates(subset=["code"]).reset_index(drop=True)
    return use

def pick_cnaps(df):
    df2 = df.copy()
    if df2.columns.size:
        df2.columns = [str(c).strip() for c in df2.columns]
    needed = ["BNKCODE","CLSCODE","CITYCODE","LNAME"]
    if set(needed).issubset(set(df2.columns)):
        use = df2[needed].copy()
    elif df2.shape[1] >= 4:
        use = df2.iloc[:, :4].copy(); use.columns = needed
        if df2.shape[1] > 4:
            extra = df2.iloc[:, 4:].astype(str).apply(lambda r: " ".join([x for x in r if x and x != "nan"]), axis=1)
            use["LNAME"] = use["LNAME"].astype(str).fillna("") + " " + extra
            use["LNAME"] = use["LNAME"].str.strip()
    elif df2.shape[1] == 1:
        col = df2.columns[0]
        parts = df2[col].str.split(r"[|\t,;]+", expand=True)
        if parts.shape[1] >= 4:
            use = parts.iloc[:, :4]; use.columns = needed
        else:
            parts = df2[col].str.split(r"\s+", expand=True)
            if parts.shape[1] >= 4:
                use = parts.iloc[:, :4]; use.columns = needed
            else:
                def regex_extract(s):
                    s = str(s or "")
                    m = re.search(r"(\d{12})", s)
                    tail = (s[m.end():].strip() if m else "")
                    return (m.group(1) if m else ""), "", "", tail.strip(" ,;|\t")
                use = pd.DataFrame([regex_extract(v) for v in df2[col].tolist()], columns=needed)
    else:
        use = df2.reindex(columns=range(4)).copy(); use.columns = needed
    use["BNKCODE"] = use["BNKCODE"].astype(str).str.replace(".0","", regex=False)
    use["BNKCODE"] = use["BNKCODE"].str.extract(r"(\d{12})", expand=False).fillna("")
    use["LNAME"] = use["LNAME"].astype(str).str.strip()
    use = use[(use["BNKCODE"]!="") & (use["LNAME"]!="")].drop_duplicates(subset=["BNKCODE"]).reset_index(drop=True)
    return use

# ---------------- UI 复用组件 ----------------
class ScrollableTree(ttk.Frame):
    def __init__(self, master, **kwargs):
        super().__init__(master)
        self.tree = ttk.Treeview(self, show="headings", **kwargs)
        xbar = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        ybar = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.tree.configure(xscrollcommand=xbar.set, yscrollcommand=ybar.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        ybar.grid(row=0, column=1, sticky="ns")
        xbar.grid(row=1, column=0, sticky="ew")
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

class ScrollableForm(ttk.Frame):
    def __init__(self, master):
        super().__init__(master)
        canvas = tk.Canvas(self, borderwidth=0, highlightthickness=0)
        hsb = ttk.Scrollbar(self, orient="horizontal", command=canvas.xview)
        vsb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.inner = ttk.Frame(canvas)
        self.inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.inner, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

def center_and_autosize(win, min_w=760, min_h=420, pad=24):
    win.update_idletasks()
    req_w = max(min_w, win.winfo_reqwidth() + pad)
    req_h = max(min_h, win.winfo_reqheight() + pad)
    try:
        parent = win.master.winfo_toplevel()
        px, py = parent.winfo_rootx(), parent.winfo_rooty()
        pw, ph = parent.winfo_width(), parent.winfo_height()
        x = px + max(0, (pw - req_w)//2)
        y = py + max(0, (ph - req_h)//2)
    except Exception:
        x, y = 100, 100
    win.minsize(req_w, req_h)
    win.geometry(f"{req_w}x{req_h}+{x}+{y}")

# ---------------- 选择行号弹窗 ----------------
class CodePicker(tk.Toplevel):
    def __init__(self, master, default_source="ibps", ibps_only=False):
        super().__init__(master)
        self.title("选择行号（IBPS/CNAPS）" if not ibps_only else "选择银行（IBPS）")
        self.resizable(True, True)
        self.ibps_only = ibps_only
        self.source = tk.StringVar(value="ibps")
        self.kw = tk.StringVar()
        self.selected_row = None
        top = ttk.Frame(self, padding=8); top.pack(fill="x")
        if not ibps_only:
            ttk.Label(top, text="来源：").pack(side="left")
            ttk.Radiobutton(top, text="IBPS（清算）", variable=self.source, value="ibps", command=self.search).pack(side="left")
            ttk.Radiobutton(top, text="CNAPS（大额）", variable=self.source, value="cnaps", command=self.search).pack(side="left", padx=8)
            self.source.set(default_source)
        else:
            ttk.Label(top, text="来源：IBPS（清算）").pack(side="left")
            self.source.set("ibps")
        ttk.Label(top, text="关键字：").pack(side="left", padx=8)
        ent = ttk.Entry(top, textvariable=self.kw, width=32); ent.pack(side="left"); ent.bind("<Return>", lambda e: self.search())
        ttk.Button(top, text="查询", command=self.search).pack(side="left", padx=6)

        self.stree = ScrollableTree(self, height=18); self.stree.pack(fill="both", expand=True, padx=8, pady=6)
        tree = self.stree.tree; tree["columns"] = ["code","name"]
        for c, w in [("code",180),("name",480)]:
            tree.heading(c, text=c); tree.column(c, width=w, anchor="w")
        tree.bind("<Double-1>", lambda e: self.pick()); tree.bind("<Return>", lambda e: self.pick())

        btns = ttk.Frame(self, padding=8); btns.pack(fill="x")
        ttk.Button(btns, text="确定", command=self.pick).pack(side="right", padx=6)
        ttk.Button(btns, text="取消", command=self.destroy).pack(side="right")

        self.after(10, lambda: (self.search(), center_and_autosize(self, 760, 520)))

    def search(self):
        rows = query(DB_PATH, self.source.get(), self.kw.get().strip(), limit=5000)
        tree = self.stree.tree; tree.delete(*tree.get_children())
        for r in rows:
            tree.insert("", "end", values=[r["code"], r["name"]])

    def pick(self):
        tree = self.stree.tree; sel = tree.selection()
        if not sel: messagebox.showinfo("提示","请先选择一条"); return
        vals = tree.item(sel[0], "values")
        self.selected_row = (vals[0], vals[1])
        self.destroy()

# ---------------- 库维护 Tab ----------------
class LibraryTab(ttk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.table_choice = tk.StringVar(value="ibps")
        self.kw = tk.StringVar()
        self._build()

    def _build(self):
        top = ttk.Frame(self); top.pack(fill="x", padx=8, pady=8)
        ttk.Label(top, text="维护库：").pack(side="left")
        ttk.Radiobutton(top, text="IBPS（清算）", variable=self.table_choice, value="ibps").pack(side="left")
        ttk.Radiobutton(top, text="CNAPS（大额）", variable=self.table_choice, value="cnaps").pack(side="left", padx=8)
        ttk.Button(top, text="导入行号", command=self.import_file).pack(side="left", padx=12)
        ttk.Button(top, text="导出库", command=self.export_db).pack(side="left", padx=12)
        ttk.Label(top, text="关键词：").pack(side="left", padx=12)
        ttk.Entry(top, textvariable=self.kw, width=28).pack(side="left")
        ttk.Button(top, text="查询", command=self.search).pack(side="left", padx=6)

        self.stree = ScrollableTree(self); self.stree.pack(fill="both", expand=True, padx=8, pady=6)
        # 背景
        try:
            _install_watermark(self, str(APP_DIR / "bg.jpg"), opacity=0.08)
        except Exception:
            pass

    def import_file(self):
        path = filedialog.askopenfilename(filetypes=[("TXT/Excel/CSV","*.txt;*.dat;*.xls;*.xlsx;*.csv"),("所有文件","*.*")])
        if not path: return
        try:
            df = read_any(path)
        except Exception as e:
            messagebox.showerror("失败", f"读取失败：{e}"); return
        rows = []; raw_src = os.path.basename(path)
        if self.table_choice.get()=="cnaps":
            use = pick_cnaps(df)
            for _, r in use.iterrows():
                code = r.get("BNKCODE",""); name = r.get("LNAME","")
                if re.fullmatch(r"\d{12}", str(code) or ""):
                    raw = "|".join([str(r.get(c,"")) for c in ["BNKCODE","CLSCODE","CITYCODE","LNAME"]])
                    rows.append((str(code), str(name), raw, raw_src))
            table="cnaps"
        else:
            use = pick_ibps(df)
            for _, r in use.iterrows():
                code = r.get("code",""); name = r.get("name","")
                if re.fullmatch(r"\d{12}", str(code) or ""):
                    raw = "|".join([str(r.get(c,"")) for c in ["code","name"]])
                    rows.append((str(code), str(name), raw, raw_src))
            table="ibps"
        if not rows:
            messagebox.showwarning("提示","未发现有效的12位行号记录（请检查文件内容/编码/格式）"); return
        if messagebox.askyesno("导入方式", "选择“是”= 全量替换；“否”= 增量合并（按 code upsert）"):
            replace_all(DB_PATH, table, rows)
        else:
            upsert_many_batched(DB_PATH, table, rows, batch_size=20000)
        messagebox.showinfo("成功", f"导入完成：共 {len(rows)} 条")

    def search(self):
        table = self.table_choice.get()
        rows = query(DB_PATH, table, self.kw.get().strip(), limit=5000)
        import pandas as pd
        df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["code","name"])
        self._load_df(df)

    def export_db(self):
        rows = query(DB_PATH, self.table_choice.get(), "", limit=999999)
        if not rows:
            messagebox.showinfo("提示","当前库为空"); return
        import pandas as pd
        df = pd.DataFrame(rows)
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel",".xlsx"),("CSV",".csv")])
        if not path: return
        try:
            if path.lower().endswith(".csv"):
                df.to_csv(path, index=False, encoding="utf-8-sig")
            else:
                export_text_xlsx(df, path, include_header=True)
            messagebox.showinfo("成功", f"已导出：{os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("失败", f"导出失败：{e}")

    def _load_df(self, df):
        tree = self.stree.tree
        tree["columns"] = list(df.columns)
        for col in df.columns:
            tree.heading(col, text=col)
            tree.column(col, width=220 if col=="name" else 160, anchor="w")
        tree.delete(*tree.get_children())
        batch = []
        for _, row in df.iterrows():
            batch.append([str(row.get(c,"")) for c in df.columns])
            if len(batch) >= 2000:
                for vals in batch: tree.insert("", "end", values=vals)
                batch.clear(); tree.update_idletasks()
        for vals in batch:
            tree.insert("", "end", values=vals)

# ---------------- 代发工资 Tab ----------------
class PayrollDialog(tk.Toplevel):
    COLS = ["收款人银行名称","收款人卡号","收款人名称","金额"]
    def __init__(self, master, init_values=None):
        super().__init__(master)
        self.title("新增/编辑 - 代发工资"); self.resizable(True, True)
        self.values = {}
        sf = ScrollableForm(self); sf.pack(fill="both", expand=True)
        frm = sf.inner
        self.vars = {}
        ttk.Label(frm, text="收款人银行名称：").grid(row=0, column=0, sticky="e", padx=6, pady=6)
        v_bank = tk.StringVar(value=(init_values.get("收款人银行名称","") if init_values else ""))
        wrap = ttk.Frame(frm)
        ttk.Entry(wrap, textvariable=v_bank, width=36).pack(side="left")
        def choose_name():
            dlg = CodePicker(self, default_source="ibps", ibps_only=True)
            self.wait_window(dlg)
            if getattr(dlg, "selected_row", None):
                _, name = dlg.selected_row
                v_bank.set(name or v_bank.get())
        ttk.Button(wrap, text="选择…", command=choose_name).pack(side="left", padx=6)
        wrap.grid(row=0, column=1, sticky="w", padx=6, pady=6)
        ttk.Label(frm, text="（可点击“选择…”从 IBPS 库带出银行名称）", foreground="#666").grid(row=0, column=2, sticky="w", padx=6)
        self.vars["收款人银行名称"] = v_bank
        labels = ["收款人卡号","收款人名称","金额"]
        for i, col in enumerate(labels, start=1):
            ttk.Label(frm, text=col + "：").grid(row=i, column=0, sticky="e", padx=6, pady=6)
            var = tk.StringVar(value=(init_values.get(col,"") if init_values else ""))
            ttk.Entry(frm, textvariable=var, width=36).grid(row=i, column=1, sticky="w", padx=6, pady=6)
            self.vars[col] = var
        btns = ttk.Frame(self, padding=8); btns.pack(fill="x")
        ttk.Button(btns, text="确定", command=self.ok).pack(side="right", padx=8)
        ttk.Button(btns, text="取消", command=self.destroy).pack(side="right")
        self.bind("<Return>", lambda e: self.ok())
        self.after(10, lambda: center_and_autosize(self, min_w=760, min_h=420))

    def ok(self):
        vals = {k: v.get().strip() for k, v in self.vars.items()}
        probs = []
        if not vals["收款人银行名称"]: probs.append("收款人银行名称 不能为空")
        if not re.fullmatch(r"\d{6,32}", vals["收款人卡号"]): probs.append("收款人卡号 必须为6-32位数字")
        try:
            if float(vals["金额"]) <= 0: probs.append("金额 必须大于0")
        except Exception:
            probs.append("金额 不是有效数字")
        if not vals["收款人名称"]: probs.append("收款人名称 不能为空")
        if probs:
            messagebox.showwarning("校验不通过", "；".join(probs)); return
        self.values = vals; self.destroy()

class PayrollTab(ttk.Frame):
    COLS = ["收款人银行名称","收款人卡号","收款人名称","金额"]
    def __init__(self, master):
        super().__init__(master)
        self.df = pd.DataFrame(columns=self.COLS)
        self._build()

    def _build(self):
        top = ttk.Frame(self); top.pack(fill="x", padx=8, pady=8)
        ttk.Button(top, text="新增", command=self.add_one).pack(side="left")
        ttk.Button(top, text="编辑选中", command=self.edit_one).pack(side="left", padx=6)
        ttk.Button(top, text="删除选中", command=self.delete_selected).pack(side="left", padx=6)
        ttk.Button(top, text="导入代发工资文件", command=self.import_file).pack(side="left", padx=12)
        ttk.Button(top, text="校验并导出（无表头）", command=self.validate_export).pack(side="left", padx=6)
        self.stree = ScrollableTree(self, height=18); self.stree.pack(fill="both", expand=True, padx=8, pady=6)
        try:
            _install_watermark(self, str(APP_DIR / "bg.jpg"), opacity=0.08)
        except Exception:
            pass
        self._reload()

    def _reload(self):
        tree = self.stree.tree
        df = self.df.fillna("")
        tree["columns"] = list(self.COLS)
        for col in self.COLS:
            tree.heading(col, text=col)
            width = 240 if "银行名称" in col else 180
            tree.column(col, width=width, anchor="w")
        tree.delete(*tree.get_children())
        batch = []
        for _, row in df.iterrows():
            batch.append([str(row.get(c, "")) for c in self.COLS])
            if len(batch) >= 1000:
                for vals in batch: tree.insert("", "end", values=vals)
                batch.clear(); tree.update_idletasks()
        for vals in batch:
            tree.insert("", "end", values=vals)

    def import_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel/CSV","*.xlsx;*.xls;*.csv"), ("所有文件","*.*")])
        if not path: return
        try:
            df = read_any(path)
        except Exception as e:
            messagebox.showerror("失败", f"读取失败：{e}"); return

        cols = [str(c).strip().replace("\ufeff","") for c in list(df.columns)]
        if set(self.COLS).issubset(set(cols)):
            df = df[self.COLS].copy()
        else:
            df = df.iloc[:, :4].copy()
            df.columns = self.COLS

        def _strip(s):
            return "" if pd.isna(s) else str(s).replace("\u3000"," ").strip()
        for c in self.COLS:
            df[c] = df[c].map(_strip)
        df = df[~(df[self.COLS].apply(lambda r: all(x=="" for x in r), axis=1))].reset_index(drop=True)

        warnings = []; bad_rows = []
        for idx, r in df.iterrows():
            ok = True
            if r["收款人银行名称"] == "": ok = False; warnings.append(f"第{idx+1}行 银行名称为空")
            if not re.fullmatch(r"\d{6,32}", r["收款人卡号"]): ok = False; warnings.append(f"第{idx+1}行 卡号非6-32位数字")
            try:
                if float(r["金额"]) <= 0: ok = False; warnings.append(f"第{idx+1}行 金额≤0")
            except Exception:
                ok = False; warnings.append(f"第{idx+1}行 金额非数字")
            if r["收款人名称"] == "": ok = False; warnings.append(f"第{idx+1}行 收款人名称为空")
            if not ok: bad_rows.append(idx)
        if bad_rows:
            df = df.drop(index=bad_rows).reset_index(drop=True)

        self.df = df
        self._reload()

        total = len(df) + len(bad_rows)
        shown = len(self.df)
        msg = f"导入处理完成：源行数 {total}，有效 {shown} 行"
        if bad_rows:
            msg += f"；已跳过 {len(bad_rows)} 行（建议在源文件修正后再导）"
        messagebox.showinfo("成功", msg)

    def add_one(self):
        dlg = PayrollDialog(self); self.wait_window(dlg)
        if getattr(dlg, "values", None):
            self.df.loc[len(self.df)] = [dlg.values.get(c,"") for c in self.COLS]; self._reload()

    def edit_one(self):
        tree = self.stree.tree; sel = tree.selection()
        if not sel: messagebox.showinfo("提示","请先选择一行"); return
        idx = tree.index(sel[0]); init = {c: str(self.df.iloc[idx][c]) for c in self.COLS}
        dlg = PayrollDialog(self, init_values=init); self.wait_window(dlg)
        if getattr(dlg, "values", None):
            for c in self.COLS: self.df.at[self.df.index[idx], c] = dlg.values.get(c,"")
            self._reload()

    def delete_selected(self):
        tree = self.stree.tree; sel = tree.selection()
        if not sel: return
        idx = tree.index(sel[0]); self.df = self.df.drop(self.df.index[idx]).reset_index(drop=True); self._reload()

    def validate_export(self):
        df = self.df.fillna("")
        probs = []
        if df["收款人银行名称"].str.strip().eq("").any(): probs.append("存在 银行名称 为空的记录")
        if (~df["收款人卡号"].astype(str).str.match(r"^\d{6,32}$", na=False)).any(): probs.append("存在 卡号 非6-32位数字")
        try:
            if (pd.to_numeric(df["金额"], errors="coerce")<=0).any(): probs.append("存在 金额≤0 或非数字")
        except Exception:
            probs.append("金额列解析异常")
        if probs: messagebox.showwarning("校验结果","；".join(probs))
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel",".xlsx")])
        if not path: return
        export_text_xlsx(self.df, path, include_header=False)
        messagebox.showinfo("成功","已导出（无表头，文本格式）")

# ---------------- 批量转账 Tab ----------------
class TransferDialog(tk.Toplevel):
    COLS = ["收款方账号","收款方户名","金额","转账方式","行别信息类型",
            "收款方银行名称","收款方银行大额支付行号/跨行清算行号","用途","明细标注"]
    def __init__(self, master, init_values=None):
        super().__init__(master)
        self.title("新增/编辑 - 批量转账"); self.resizable(True, True)
        self.values = {}
        self.transfer_mode_map = {"0":"0 - 行内转账（同一银行）","1":"1 - 跨行转账（不同银行）"}
        self.bankinfo_type_map = {"":"（空）不填","0":"0 - 跨行清算银行信息（IBPS）","1":"1 - 开户支行信息（开户支行名称/大额）"}
        container = ttk.Frame(self, padding=12); container.pack(fill="both", expand=True)
        container.grid_columnconfigure(1, weight=1)
        tips = {
            "转账方式":"0=行内；1=跨行（跨行需填写下方“行号”）",
            "行别信息类型":"可为空；0=跨行清算银行信息；1=开户支行信息。行内不填；银联卡号可不填。",
            "行号":"支持手输或点击【选择…】从本地库选；选择后会自动带出银行名称，并根据是否“华夏银行”自动设置转账方式。",
            "可选项":"行内转账时不输入；收款方账号是银联卡号时可以不输入"
        }
        self.vars = {}; row=0
        def add(label, widget, hint_key=None, below_hint_key=None):
            nonlocal row
            ttk.Label(container, text=label+"：").grid(row=row, column=0, sticky="e", padx=6, pady=6)
            w = widget(container)
            if isinstance(w, tuple):
                main, extra = w; main.grid(row=row, column=1, sticky="ew", padx=6, pady=6)
                if extra is not None: extra.grid(row=row, column=2, sticky="w", padx=6, pady=6)
            else:
                w.grid(row=row, column=1, sticky="ew", padx=6, pady=6)
            if hint_key:
                ttk.Label(container, text="（"+tips[hint_key]+"）", foreground="#666").grid(row=row, column=3, sticky="w", padx=6, pady=6)
            row+=1
            if below_hint_key:
                ttk.Label(container, text="（"+tips[below_hint_key]+"）", foreground="#666").grid(row=row, column=1, columnspan=3, sticky="w", padx=6, pady=(0,8))
                row+=1
        v_acct = tk.StringVar(value=(init_values.get("收款方账号","") if init_values else "")); add("收款方账号", lambda p: ttk.Entry(p, textvariable=v_acct)); self.vars["收款方账号"]=v_acct
        v_uname= tk.StringVar(value=(init_values.get("收款方户名","") if init_values else "")); add("收款方户名", lambda p: ttk.Entry(p, textvariable=v_uname)); self.vars["收款方户名"]=v_uname
        v_amt  = tk.StringVar(value=(init_values.get("金额","") if init_values else "")); add("金额", lambda p: ttk.Entry(p, textvariable=v_amt)); self.vars["金额"]=v_amt
        cur_mode = str(init_values.get("转账方式","")) if init_values else "0"
        if cur_mode not in ("0","1"): cur_mode="0"
        v_mode = tk.StringVar(value=self.transfer_mode_map[cur_mode])
        add("转账方式", lambda p: ttk.Combobox(p, textvariable=v_mode, state="readonly", values=[self.transfer_mode_map["0"], self.transfer_mode_map["1"]]), "转账方式"); self.vars["转账方式"]=v_mode
        cur_btype = str(init_values.get("行别信息类型","")) if init_values else ""
        if cur_btype not in ("","0","1"): cur_btype=""
        v_btype= tk.StringVar(value=self.bankinfo_type_map[cur_btype])
        add("行别信息类型", lambda p: ttk.Combobox(p, textvariable=v_btype, state="readonly", values=[self.bankinfo_type_map[""], self.bankinfo_type_map["0"], self.bankinfo_type_map["1"]]), "行别信息类型"); self.vars["行别信息类型"]=v_btype
        v_bname= tk.StringVar(value=(init_values.get("收款方银行名称","") if init_values else ""))
        add("收款方银行名称", lambda p: ttk.Entry(p, textvariable=v_bname), "可选项"); self.vars["收款方银行名称"]=v_bname
        v_code = tk.StringVar(value=(init_values.get("收款方银行大额支付行号/跨行清算行号","") if init_values else ""))
        def code_builder(p):
            wrapper = ttk.Frame(p); ttk.Entry(wrapper, textvariable=v_code, width=30).pack(side="left")
            def do_pick():
                dlg = CodePicker(self)  # 双库可选
                self.wait_window(dlg)
                if getattr(dlg, "selected_row", None):
                    code, name = dlg.selected_row
                    v_code.set(code)
                    self.vars["收款方银行名称"].set(name or self.vars["收款方银行名称"].get())
                    if "华夏银行" in (name or ""):
                        self.vars["转账方式"].set(self.transfer_mode_map["0"])
                        self.vars["行别信息类型"].set(self.bankinfo_type_map[""])
                    else:
                        self.vars["转账方式"].set(self.transfer_mode_map["1"])
            ttk.Button(wrapper, text="选择…", command=do_pick).pack(side="left", padx=6)
            return (wrapper, None)
        add("收款方银行大额支付行号/跨行清算行号", code_builder, "行号", "可选项"); self.vars["收款方银行大额支付行号/跨行清算行号"]=v_code
        v_use  = tk.StringVar(value=(init_values.get("用途","") if init_values else "")); add("用途", lambda p: ttk.Entry(p, textvariable=v_use)); self.vars["用途"]=v_use
        v_note = tk.StringVar(value=(init_values.get("明细标注","") if init_values else "")); add("明细标注", lambda p: ttk.Entry(p, textvariable=v_note)); self.vars["明细标注"]=v_note
        btns = ttk.Frame(self, padding=8); btns.pack(fill="x")
        ttk.Button(btns, text="确定", command=self.ok).pack(side="right", padx=8)
        ttk.Button(btns, text="取消", command=self.destroy).pack(side="right")
        self.bind("<Return>", lambda e: self.ok())
        self.after(10, lambda: center_and_autosize(self, min_w=880, min_h=560))

    def ok(self):
        inverse_mode = {v:k for k,v in self.transfer_mode_map.items()}
        inverse_btype= {v:k for k,v in self.bankinfo_type_map.items()}
        v={}
        for k, var in self.vars.items():
            val = var.get().strip()
            if k=="转账方式": val = inverse_mode.get(val, "0")
            if k=="行别信息类型": val = inverse_btype.get(val, "")  # 允许空
            v[k]=val
        probs=[]
        if not v["收款方户名"]: probs.append("收款方户名 不能为空")
        if not re.fullmatch(r"\d{6,32}", v["收款方账号"]): probs.append("收款方账号 必须为6-32位数字")
        try:
            if float(v["金额"]) <= 0: probs.append("金额 必须大于0")
        except Exception:
            probs.append("金额 不是有效数字")
        if v.get("转账方式","")=="1" and not v.get("收款方银行大额支付行号/跨行清算行号",""):
            probs.append("跨行转账时需提供行号")
        if probs: messagebox.showwarning("校验不通过","；".join(probs)); return
        self.values = v; self.destroy()

class TransferTab(ttk.Frame):
    COLS = ["收款方账号","收款方户名","金额","转账方式","行别信息类型",
            "收款方银行名称","收款方银行大额支付行号/跨行清算行号","用途","明细标注"]
    def __init__(self, master):
        super().__init__(master)
        self.df = pd.DataFrame(columns=self.COLS); self._build()
    def _build(self):
        top = ttk.Frame(self); top.pack(fill="x", padx=8, pady=8)
        ttk.Button(top, text="新增", command=self.add_one).pack(side="left")
        ttk.Button(top, text="编辑选中", command=self.edit_one).pack(side="left", padx=6)
        ttk.Button(top, text="删除选中", command=self.delete_selected).pack(side="left", padx=6)
        ttk.Button(top, text="导入批量转账文件", command=self.import_file).pack(side="left", padx=12)
        ttk.Button(top, text="校验并导出（保留表头）", command=self.validate_export).pack(side="left", padx=6)
        self.stree = ScrollableTree(self, height=18); self.stree.pack(fill="both", expand=True, padx=8, pady=6)
        try:
            _install_watermark(self, str(APP_DIR / "bg.jpg"), opacity=0.08)
        except Exception:
            pass
        self._reload()
    def _reload(self):
        tree = self.stree.tree
        df = self.df
        tree["columns"] = list(df.columns)
        for col in df.columns:
            tree.heading(col, text=col)
            width = 220 if ("名称" in col or "用途" in col or "明细" in col) else 160
            tree.column(col, width=width, anchor="w")
        tree.delete(*tree.get_children())
        batch = []
        for _, row in df.iterrows():
            batch.append([str(row.get(c,"")) for c in df.columns])
            if len(batch) >= 1000:
                for vals in batch: tree.insert("", "end", values=vals)
                batch.clear(); tree.update_idletasks()
        for vals in batch:
            tree.insert("", "end", values=vals)
    def add_one(self):
        dlg = TransferDialog(self); self.wait_window(dlg)
        if getattr(dlg, "values", None):
            self.df.loc[len(self.df)] = [dlg.values.get(c,"") for c in self.COLS]; self._reload()
    def edit_one(self):
        tree = self.stree.tree; sel = tree.selection()
        if not sel: messagebox.showinfo("提示","请先选择一行"); return
        idx = tree.index(sel[0]); init = {c: str(self.df.iloc[idx][c]) for c in self.COLS}
        dlg = TransferDialog(self, init_values=init); self.wait_window(dlg)
        if getattr(dlg, "values", None):
            for c in self.COLS: self.df.at[self.df.index[idx], c] = dlg.values.get(c,"")
            self._reload()
    def delete_selected(self):
        tree = self.stree.tree; sel = tree.selection()
        if not sel: return
        idx = tree.index(sel[0]); self.df = self.df.drop(self.df.index[idx]).reset_index(drop=True); self._reload()
    def import_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel/CSV","*.xlsx;*.xls;*.csv")])
        if not path: return
        try:
            df = read_any(path)
        except Exception as e:
            messagebox.showerror("失败", f"读取失败：{e}"); return
        need = self.COLS
        if not set(need).issubset(set(df.columns)):
            df = df.iloc[:, :9]
            df.columns = need
        errors = []
        for idx, r in df.fillna("").iterrows():
            if not re.fullmatch(r"\d{6,32}", str(r["收款方账号"])): errors.append(f"第{idx+1}行：收款方账号 非6-32位数字")
            if str(r["收款方户名"]).strip()=="" : errors.append(f"第{idx+1}行：收款方户名 为空")
            try:
                if float(str(r["金额"]))<=0: errors.append(f"第{idx+1}行：金额 ≤ 0")
            except Exception:
                errors.append(f"第{idx+1}行：金额 非数字")
            mode = str(r["转账方式"]).strip()
            if mode not in ("0","1"): errors.append(f"第{idx+1}行：转账方式 非 0/1")
            btype = str(r["行别信息类型"]).strip()
            if btype not in ("","0","1"): errors.append(f"第{idx+1}行：行别信息类型 只能为空/0/1")
            if mode=="1" and str(r["收款方银行大额支付行号/跨行清算行号"]).strip()=="":
                errors.append(f"第{idx+1}行：跨行转账需提供行号")
        if errors:
            messagebox.showerror("校验失败","导入中止：\\n" + "\\n".join(errors[:30]) + ("\\n..." if len(errors)>30 else ""))
            return
        self.df = df[self.COLS].astype(str); self._reload()
        messagebox.showinfo("成功","导入成功，已加载到下方明细，可继续编辑。")
    def validate_export(self):
        df = self.df.fillna("")
        probs = []
        if (~df["收款方账号"].astype(str).str.match(r"^\\d{6,32}$", na=False)).any(): probs.append("收款方账号 格式异常")
        if df["收款方户名"].str.strip().eq("").any(): probs.append("收款方户名 不能为空")
        try:
            if (pd.to_numeric(df["金额"], errors="coerce")<=0).any(): probs.append("金额 必须大于0")
        except Exception:
            probs.append("金额 不是可解析数字")
        cross = df["转账方式"].astype(str).str.strip()=="1"
        need = cross & df["收款方银行大额支付行号/跨行清算行号"].astype(str).str.strip().eq("")
        if need.any(): probs.append("跨行转账时需提供行号（IBPS或CNAPS）")
        if probs: messagebox.showwarning("校验结果","；".join(probs))
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel",".xlsx")])
        if not path: return
        export_text_xlsx(self.df, path, include_header=True)
        messagebox.showinfo("成功","已导出（保留表头，文本格式）")

# ---------------- 帮助菜单：环境自检 ----------------
def show_env_check():
    msgs = []
    ok = True
    try:
        import pandas as _p
        msgs.append(f"pandas: {_p.__version__}")
    except Exception as e:
        ok=False; msgs.append(f"pandas: 未安装 ({e})")
    try:
        import openpyxl as _o
        msgs.append(f"openpyxl: {_o.__version__}")
    except Exception as e:
        ok=False; msgs.append(f"openpyxl: 未安装 ({e})")
    try:
        import xlrd as _x
        v = getattr(_x,'__version__','?')
        msgs.append(f"xlrd: {v} (需要 1.2.0 以读取 .xls)")
        if v != "1.2.0":
            ok=False
    except Exception as e:
        ok=False; msgs.append(f"xlrd: 未安装 ({e})")
    guide = ""
    if not ok:
        guide = "\\n\\n修复指引（在命令行执行）：\\n" + \
                "pip uninstall -y xlrd\\n" + \
                "pip install xlrd==1.2.0\\n" + \
                "pip install openpyxl==3.1.2 pandas==2.2.2"
    messagebox.showinfo("环境自检", "\\n".join(msgs) + guide)

# ---------------- 应用主窗体 ----------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("华夏离线批量编辑器 v2.3.6-r3")
        self.minsize(1200, 760)
        try:
            self.iconbitmap(str(APP_DIR / "icon.ico"))
        except Exception:
            pass
        ensure_db(DB_PATH)

        # 菜单
        menubar = tk.Menu(self)
        viewm = tk.Menu(menubar, tearoff=0)
        def _choose_bg():
            p = filedialog.askopenfilename(title='选择背景图片', filetypes=[('图片','*.jpg;*.jpeg;*.png;*.webp;*.bmp')])
            if not p: return
            for tab in self.notebook_tabs:
                _install_watermark(tab, p, opacity=getattr(self, "_bg_opacity", 0.08))
        def _opacity(val=None):
            try:
                self._bg_opacity = float(val)
            except Exception:
                self._bg_opacity = 0.08
            for tab in self.notebook_tabs:
                _install_watermark(tab, str(APP_DIR / "bg.jpg"), opacity=self._bg_opacity)
        viewm.add_command(label='设置背景图…', command=_choose_bg)
        viewm.add_separator()
        viewm.add_radiobutton(label='透明度 5%',  command=lambda: _opacity(0.05))
        viewm.add_radiobutton(label='透明度 8%',  command=lambda: _opacity(0.08))
        viewm.add_radiobutton(label='透明度 12%', command=lambda: _opacity(0.12))
        menubar.add_cascade(label='视图', menu=viewm)

        helpm = tk.Menu(menubar, tearoff=0)
        helpm.add_command(label="环境自检与修复…", command=show_env_check)
        helpm.add_separator()
        helpm.add_command(label="关于", command=lambda: messagebox.showinfo("关于","华夏离线批量编辑器 v2.3.6-r3"))
        menubar.add_cascade(label="帮助", menu=helpm)
        self.config(menu=menubar)

        nb = ttk.Notebook(self); nb.pack(fill="both", expand=True)
        t1 = LibraryTab(nb); t2 = PayrollTab(nb); t3 = TransferTab(nb)
        nb.add(t1, text="库维护（IBPS/CNAPS 导入）")
        nb.add(t2, text="代发工资录入")
        nb.add(t3, text="批量转账录入")
        self.notebook_tabs = [t1, t2, t3]

        # 初始加载根目录的 bg.jpg
        for tab in self.notebook_tabs:
            try:
                _install_watermark(tab, str(APP_DIR / "bg.jpg"), opacity=getattr(self, "_bg_opacity", 0.08))
            except Exception:
                pass

if __name__ == "__main__":
    App().mainloop()
